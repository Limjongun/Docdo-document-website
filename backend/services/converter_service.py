import io
import fitz  # PyMuPDF
import pdfplumber
import openpyxl
from PIL import Image
from docx import Document
from openpyxl import Workbook


def images_to_pdf(image_bytes_list: list[bytes]) -> bytes:
    """Merge multiple images into a single PDF."""
    pil_images = []
    for b in image_bytes_list:
        img = Image.open(io.BytesIO(b)).convert("RGB")
        pil_images.append(img)

    if not pil_images:
        raise ValueError("No images provided")

    buf = io.BytesIO()
    pil_images[0].save(
        buf,
        format="PDF",
        save_all=True,
        append_images=pil_images[1:],
    )
    return buf.getvalue()


def pdf_to_excel(pdf_bytes: bytes) -> bytes:
    """Extract tables from PDF and write to .xlsx."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            ws = wb.create_sheet(title=f"Page {page_num}")
            tables = page.extract_tables()
            if tables:
                row_idx = 1
                for table in tables:
                    for row in table:
                        for col_idx, cell in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell or "")
                        row_idx += 1
                    row_idx += 1  # blank row between tables
            else:
                # Fallback: dump raw text
                ws.cell(row=1, column=1, value=page.extract_text() or "")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_to_pdf(excel_bytes: bytes) -> bytes:
    """Convert Excel sheets to PDF by rendering as text (basic conversion)."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    doc = fitz.open()

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        page = doc.new_page(width=595, height=842)  # A4
        text_lines = [f"Sheet: {sheet}\n"]
        for row in ws.iter_rows(values_only=True):
            line = "\t".join(str(c) if c is not None else "" for c in row)
            text_lines.append(line)

        page.insert_text(
            point=(30, 50),
            text="\n".join(text_lines),
            fontsize=9,
        )

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()
